import openpyxl

book = openpyxl.load_workbook("C:\\Users\\chand\\OneDrive\\Desktop\\PS\\Data.xlsx")
sheet = book.active


dataDict = {}

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TC2":
        for j in range(2, sheet.max_column+1):
            dataDict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        print(dataDict)
