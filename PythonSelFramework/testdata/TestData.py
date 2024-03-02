import openpyxl

class TestDataHomePage:
    test_homepage = [{"name":"chandani","email":"chandani.jain@gmail.com", "pwd":"abc"},{"name":"deepak", "email":"deepak.kothari@gmail.com", "pwd":"xyz"}]

    @staticmethod
    def getTestData(testcase):
        book = openpyxl.load_workbook("C:\\Users\\chand\\OneDrive\\Desktop\\PS\\Data.xlsx")
        sheet = book.active

        dataDict = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase:
                for j in range(2, sheet.max_column + 1):
                    dataDict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dataDict]